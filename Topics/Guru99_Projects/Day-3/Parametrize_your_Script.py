import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
import pandas as pd
import openpyxl

options = Options()
options.headless = False
serviceObj = Service("/Users/mac/PycharmProjects/Selenium_practice/Topics/driver/chromedriver")
driver = webdriver.Chrome(service=serviceObj, options=options)


def readValues():
    file = '/Users/mac/Downloads/utilities.xlsx'

    book = openpyxl.load_workbook(file)
    sheet = book.active

    user_id = sheet.cell(row=2, column=1).value
    password = sheet.cell(row=2, column=2).value
    return user_id, password


def goToSite():
    driver.get("https://www.demo.guru99.com/V4/index.php")


def validUserId_password(user_id, pas):
    goToSite()
    driver.implicitly_wait(10)
    user_id_input = 'input[type="text"]'
    password_input = 'input[type="password"]'
    login_btn = '[type="submit"]'
    driver.find_element(By.CSS_SELECTOR, user_id_input).send_keys(user_id)  # Step-2: Enter Valid User Id
    driver.find_element(By.CSS_SELECTOR, password_input).send_keys(pas)  # Step-3: Enter Valid Password
    driver.implicitly_wait(10)

    driver.find_element(By.CSS_SELECTOR, login_btn).click()  # Step-4: Click Login
    driver.implicitly_wait(10)

    assert "Guru99 Bank Manager HomePage" in driver.title

    if driver.title == "Guru99 Bank Manager HomePage":
        print("Login Successfull")

    time.sleep(2)
    driver.quit()


us, pas = readValues()

validUserId_password(us, pas)
